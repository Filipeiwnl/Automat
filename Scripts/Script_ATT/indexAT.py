import os
import requests
from bs4 import BeautifulSoup
import pandas as pd
import re
from openpyxl import load_workbook
from datetime import datetime

def get_csrf_token(session, url):
    response = session.get(url)
    if response.status_code == 200:
        soup = BeautifulSoup(response.content, 'html.parser')
        csrf_token = soup.find('input', {'name': 'csrfmiddlewaretoken'})['value']
        return csrf_token
    else:
        raise Exception(f"Failed to get CSRF token. Status code: {response.status_code}")

def extract_span_info(info_str):
    info_dict = {}
    matches = re.findall(r'(\w+):([^\n]+)', info_str)
    for match in matches:
        key, value = match
        info_dict[key.strip()] = value.strip()
    return info_dict

def trat_niveis(info_str):
    falhas = ['S/ GER', 'S/ PLC', 'LOS', 'S/ IP']
    matches = re.findall(r'(\w+):([^\n]+)', info_str)
    trat_dict = {}
    for match in matches:
        key, value = match
        value = value.strip()
        if value not in falhas:
            trat_dict[key.strip()] = 'sem Falhas'
        else:
            trat_dict[key.strip()] = value
    return trat_dict

def safe_float_conversion(value):
    try:
        return float(value)
    except ValueError:
        return 0.0

def consultar_dados(session, base_url, designador):
    consulta_url = f"{base_url}/dados_enlace/{designador}"
    print(f"Consultando URL: {consulta_url}")

    response = session.get(consulta_url)
    if response.status_code != 200:
        print(f"Falha na consulta. Status code: {response.status_code}")
        return None
    
    soup = BeautifulSoup(response.content, 'html.parser')
    
    t_pontas = soup.find_all('div', class_='tPontas')
    if len(t_pontas) >= 2:
        ponta_a = t_pontas[0].text.strip()
        ponta_b = t_pontas[1].text.strip()
    else:
        ponta_a = ponta_b = 'N/A'
    print(f"Ponta A: {ponta_a}, Ponta B: {ponta_b}")
    
    csrf_token = get_csrf_token(session, consulta_url)

    data = {
        'id_ots': designador,
        'is_secure': 'true',  
        'csrfmiddlewaretoken': csrf_token
    }

    headers = {
        'Referer': consulta_url
    }

    response = session.post(f"{base_url}/ajax_span", data=data, headers=headers)
    if response.status_code == 200:
        response_json = response.json()
        
        print("Response JSON:", response_json)
        
        comandoAjaxSpan = response_json.get('comandoAjaxSpan', {})
        link_a_info = extract_span_info(comandoAjaxSpan.get('link_a_info', ''))
        link_b_info = extract_span_info(comandoAjaxSpan.get('link_b_info', ''))
        trat_link_a_info = trat_niveis(comandoAjaxSpan.get('link_a_info', ''))
        trat_link_b_info = trat_niveis(comandoAjaxSpan.get('link_b_info', ''))

        km = safe_float_conversion(comandoAjaxSpan.get('distancia_a_b', 0))
        tx_a = safe_float_conversion(link_a_info.get('TX', 0))
        rx_a = safe_float_conversion(link_a_info.get('RX', 0))
        tx_b = safe_float_conversion(link_b_info.get('TX', 0))
        rx_b = safe_float_conversion(link_b_info.get('RX', 0))

        if km != 0:
            db_km_a = round((rx_a - tx_a) / km, 2)
            db_km_b = round((rx_b - tx_b) / km, 2)
        else:
            db_km_a = db_km_b = 'N/A'

        dados = {
            'TX A': link_a_info.get('TX', 'N/A'),
            'RX A': link_a_info.get('RX', 'N/A'),
            'Span atual A': link_a_info.get('SPAN', 'N/A'),
            'Span projetado A': link_a_info.get('SPAN_PROJ', 'N/A'),
            'Histórico A': link_a_info.get('SPAN_HIST', 'N/A'),
            'KM': comandoAjaxSpan.get('distancia_a_b', 'N/A'),
            'RX B': link_b_info.get('RX', 'N/A'),
            'TX B': link_b_info.get('TX', 'N/A'),
            'Span atual B': link_b_info.get('SPAN', 'N/A'),
            'Span projetado B': link_b_info.get('SPAN_PROJ', 'N/A'),
            'Histórico B': link_b_info.get('SPAN_HIST', 'N/A'),
            'Ponta A': ponta_a,
            'Ponta B': ponta_b,
            'Status A': comandoAjaxSpan.get('link_a_status', 'N/A'),
            'Status B': comandoAjaxSpan.get('link_b_status', 'N/A'),
            'Trat Niveis A': trat_link_a_info.get('TX', 'N/A'),
            'Trat Niveis B': trat_link_b_info.get('TX', 'N/A'),
            'DB/KM A': db_km_a,
            'DB/KM B': db_km_b
        }
        
        print("Dados extraídos:", dados)
        
        return dados
    else:
        print(f"Falha na consulta. Status code: {response.status_code}")
        return None

def main():
    base_url = '//IP'
    session = requests.Session()
    
    file_path = os.path.join('Planilhas', 'OTSS.xlsx')
    
    os.makedirs(os.path.dirname(file_path), exist_ok=True)

    if not os.path.isfile(file_path):
        raise FileNotFoundError(f"O arquivo {file_path} não foi encontrado.")
    if not file_path.endswith('.xlsx'):
        raise ValueError(f"O arquivo {file_path} não é um arquivo Excel válido.")
    
    try:
        df_input = pd.read_excel(file_path, engine='openpyxl')
    except Exception as e:
        raise ValueError(f"Erro ao abrir o arquivo {file_path}. Verifique se é um arquivo Excel válido e não está corrompido.") from e
    
    print("Colunas do DataFrame:")
    print(df_input.columns)
    
    required_columns = ['OTS', 'TX A', 'RX A', 'Span atual A', 'Span projetado A', 'Histórico A',
                        'Equipamento de Envio (A)', 'TX B', 'RX B', 'Span atual B', 'Span projetado B',
                        'Histórico B', 'Equipamento de Recepção (B)', 'KM', 'Status A', 'Status B', 
                        'Trat Niveis A', 'Trat Niveis B', 'DB/KM A', 'DB/KM B', 'ATUALIZADO EM', 'AVISOS']
    for col in required_columns:
        if col not in df_input.columns:
            df_input[col] = ''  # Adiciona a coluna se não existir
    
    num_rows = len(df_input)

    print(f"Total de linhas no DataFrame: {num_rows}")

    for index in range(num_rows):
        row = df_input.iloc[index]

        designador = row['OTS']
        if pd.notna(designador) and str(designador).strip() != '':
            dados_extraidos = consultar_dados(session, base_url, designador)
            if dados_extraidos is not None:
                df_input.at[index, 'TX A'] = str(dados_extraidos['TX A'])
                df_input.at[index, 'RX A'] = str(dados_extraidos['RX A'])
                df_input.at[index, 'Span atual A'] = str(dados_extraidos['Span atual A'])
                df_input.at[index, 'Span projetado A'] = str(dados_extraidos['Span projetado A'])
                df_input.at[index, 'DB/KM A'] = str(dados_extraidos['DB/KM A'])
                df_input.at[index, 'Trat Niveis A'] = str(dados_extraidos['Trat Niveis A'])
                df_input.at[index, 'Status A'] = str(dados_extraidos['Status A'])
                df_input.at[index, 'TX B'] = str(dados_extraidos['TX B'])
                df_input.at[index, 'RX B'] = str(dados_extraidos['RX B'])
                df_input.at[index, 'Span atual B'] = str(dados_extraidos['Span atual B'])
                df_input.at[index, 'Span projetado B'] = str(dados_extraidos['Span projetado B'])
                df_input.at[index, 'DB/KM B'] = str(dados_extraidos['DB/KM B'])
                df_input.at[index, 'Trat Niveis B'] = str(dados_extraidos['Trat Niveis B'])
                df_input.at[index, 'Status B'] = str(dados_extraidos['Status B'])
                
                # Adiciona avisos se RX ou Span atual contiverem texto
                avisos = []
                if not safe_float_conversion(dados_extraidos['RX A']):
                    avisos.append(f"RX A: {dados_extraidos['RX A']}")
                if not safe_float_conversion(dados_extraidos['RX B']):
                    avisos.append(f"RX B: {dados_extraidos['RX B']}")
                if not safe_float_conversion(dados_extraidos['Span atual A']):
                    avisos.append(f"Span atual A: {dados_extraidos['Span atual A']}")
                if not safe_float_conversion(dados_extraidos['Span atual B']):
                    avisos.append(f"Span atual B: {dados_extraidos['Span atual B']}")
                
                df_input.at[index, 'AVISOS'] = ', '.join(avisos)
    
    # Atualiza a coluna 'ATUALIZADO EM' com a data e hora atuais
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    df_input['ATUALIZADO EM'] = timestamp

    book = load_workbook(file_path)
    sheet_name = book.sheetnames[0]
    sheet = book[sheet_name]
    
    # Limpa as células para evitar sobreposição de dados antigos com os novos
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.value = None

    for r_idx, row in df_input.iterrows():
        for c_idx, value in enumerate(row):
            sheet.cell(row=r_idx + 2, column=c_idx + 1, value=value)
    
    book.save(file_path)
    print("Consulta e gravação de resultados concluídas com sucesso.")

if __name__ == "__main__":
    main()
