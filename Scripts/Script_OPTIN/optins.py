import os
import requests
from bs4 import BeautifulSoup
import pandas as pd
import re
from openpyxl import load_workbook
from datetime import datetime
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule, FormulaRule

#### TOKEN VALIDATION ####
def get_csrf_token(session, url):
    response = session.get(url)
    if response.status_code == 200:
        soup = BeautifulSoup(response.content, 'html.parser')
        csrf_token = soup.find('input', {'name': 'csrfmiddlewaretoken'})['value']
        return csrf_token
    else:
        raise Exception(f"Failed to get CSRF token. Status code: {response.status_code}")

#### DATA OF THE PROBLEMS EXTRACTION ####
def extract_span_info(info_str):
    info_dict = {}
    matches = re.findall(r'(\w+):([^\n]+)', info_str)
    for match in matches:
        key, value = match
        info_dict[key.strip()] = value.strip()
    return info_dict

#### DATA OF THE PROBLEMS TRATMENT #### 
def trat_niveis(info_str):
    falhas = ['S/ GER', 'S/ PLC', 'LOS', 'S/ IP']
    matches = re.findall(r'(\w+):([^\n]+)', info_str)
    trat_dict = {}
    for match in matches:
        key, value = match
        value = value.strip()
        if value not in falhas:
            trat_dict[key.strip()] = 'sem Falhas'
        else:
            trat_dict[key.strip()] = value
    return trat_dict

def safe_float_conversion(value):
    try:
        return float(value)
    except ValueError:
        return 0.0

def data_consult(session, base_url, designador):
    consult_url = f"{base_url}/dados_enlace/{designador}"
    print(f"Consulting URL: {consult_url}")

    response = session.get(consult_url)
    if response.status_code != 200:
        print(f"Failed Query for URL {consult_url}. Status code: {response.status_code}")
        return None
    
    soup = BeautifulSoup(response.content, 'html.parser')
    
    t_pontas = soup.find_all('div', class_='tPontas')
    if len(t_pontas) >= 2:
        ponta_a = t_pontas[0].text.strip()
        ponta_b = t_pontas[1].text.strip()
    else:
        ponta_a = ponta_b = 'N/A'
    print(f"Ponta A: {ponta_a}, Ponta B: {ponta_b}")
    
    csrf_token = get_csrf_token(session, consult_url)

    data = {
        'id_ots': designador,
        'is_secure': 'true',
        'csrfmiddlewaretoken': csrf_token
    }

    headers = {
        'Referer': consult_url
    }

    response = session.post(f"{base_url}/ajax_span", data=data, headers=headers)
    if response.status_code == 200:
        response_json = response.json()
        
        print("Response JSON:", response_json)
        
        comandoAjaxSpan = response_json.get('comandoAjaxSpan', {})
        link_a_info = extract_span_info(comandoAjaxSpan.get('link_a_info', ''))
        link_b_info = extract_span_info(comandoAjaxSpan.get('link_b_info', ''))
        trat_link_a_info = trat_niveis(comandoAjaxSpan.get('link_a_info', ''))
        trat_link_b_info = trat_niveis(comandoAjaxSpan.get('link_b_info', ''))

        km = safe_float_conversion(comandoAjaxSpan.get('distancia_a_b', 0))
        tx_a = safe_float_conversion(link_a_info.get('TX', 0))
        rx_a = safe_float_conversion(link_a_info.get('RX', 0))
        tx_b = safe_float_conversion(link_b_info.get('TX', 0))
        rx_b = safe_float_conversion(link_b_info.get('RX', 0))

        if km != 0:
            db_km_a = round((rx_a - tx_a) / km, 2)
            db_km_b = round((rx_b - tx_b) / km, 2)
        else:
            db_km_a = db_km_b = 'N/A'

        Information = {
            'TX A': link_a_info.get('TX', 'N/A'),
            'RX A': link_a_info.get('RX', 'N/A'),
            'Span atual A': link_a_info.get('SPAN', 'N/A'),
            'Span projetado A': link_a_info.get('SPAN_PROJ', 'N/A'),
            'Histórico A': link_a_info.get('SPAN_HIST', 'N/A'),
            'KM: A <> B': comandoAjaxSpan.get('distancia_a_b', 'N/A'),
            'RX B': link_b_info.get('RX', 'N/A'),
            'TX B': link_b_info.get('TX', 'N/A'),
            'Span atual B': link_b_info.get('SPAN', 'N/A'),
            'Span projetado B': link_b_info.get('SPAN_PROJ', 'N/A'),
            'Histórico B': link_b_info.get('SPAN_HIST', 'N/A'),
            'Ponta A': ponta_a,
            'Ponta B': ponta_b,
            'Status A': comandoAjaxSpan.get('link_a_status', 'N/A'),
            'Status B': comandoAjaxSpan.get('link_b_status', 'N/A'),
            'Trat Niveis A': trat_link_a_info.get('TX', 'N/A'),
            'Trat Niveis B': trat_link_b_info.get('TX', 'N/A'),
            'DB/KM A': db_km_a,
            'DB/KM B': db_km_b
        }
        
        print("Information extraídos:", Information)
        
        return Information, consult_url  # Retorna os Information e a URL de consulta
    else:
        print(f"Failed to retrieve data. Status code: {response.status_code}")
        return None  # Retorna None em caso de falha

def is_numeric(value):
    """
    Checks if a value is numeric
    """
    try:
        float(value)
        return True
    except ValueError:
        return False

#### Update Sheet ####
def atualizar_planilha(df_input, session, base_url, designadores=None, novos_Information=False):
    num_rows = len(df_input)
    atualizados = []

    for index in range(num_rows):
        row = df_input.iloc[index]
        designador = row['OTS']
        
        if designadores:
            if designador not in designadores:
                continue
        
        if novos_Information and pd.notna(row['Equipamento de Envio (A)']) and str(row['Equipamento de Envio (A)']).strip() != '':
            print(f"Linha {index + 1} já tem um resultado, pulando a consulta.")
            continue

        if pd.notna(designador) and str(designador).strip() != '':
            result = data_consult(session, base_url, designador)
            if result:
                Information_extraidos, consult_url = result
                df_input.at[index, 'TX A'] = str(Information_extraidos['TX A'])
                df_input.at[index, 'RX A'] = str(Information_extraidos['RX A'])
                df_input.at[index, 'Span atual A'] = str(Information_extraidos['Span atual A'])
                df_input.at[index, 'Span projetado A'] = str(Information_extraidos['Span projetado A'])
                df_input.at[index, 'Histórico A'] = str(Information_extraidos['Histórico A'])
                df_input.at[index, 'Equipamento de Envio (A)'] = str(Information_extraidos['Ponta A'])
                df_input.at[index, 'TX B'] = str(Information_extraidos['TX B'])
                df_input.at[index, 'RX B'] = str(Information_extraidos['RX B'])
                df_input.at[index, 'Span atual B'] = str(Information_extraidos['Span atual B'])
                df_input.at[index, 'Span projetado B'] = str(Information_extraidos['Span projetado B'])
                df_input.at[index, 'Histórico B'] = str(Information_extraidos['Histórico B'])
                df_input.at[index, 'Equipamento de Recepção (B)'] = str(Information_extraidos['Ponta B'])
                df_input.at[index, 'KM: A <> B'] = str(Information_extraidos['KM: A <> B'])
                df_input.at[index, 'Status A'] = str(Information_extraidos['Status A'])
                df_input.at[index, 'Status B'] = str(Information_extraidos['Status B'])
                df_input.at[index, 'Trat Niveis A'] = str(Information_extraidos['Trat Niveis A'])
                df_input.at[index, 'Trat Niveis B'] = str(Information_extraidos['Trat Niveis B'])
                df_input.at[index, 'DB/KM A'] = str(Information_extraidos['DB/KM A'])
                df_input.at[index, 'DB/KM B'] = str(Information_extraidos['DB/KM B'])
                
                # Add "OK" to "AVISOS" if there are no issues
                avisos = []
                
                if not is_numeric(Information_extraidos['RX A']):
                    avisos.append(f"{Information_extraidos['RX A']}")
                if not is_numeric(Information_extraidos['RX B']):
                    avisos.append(f"{Information_extraidos['RX B']}")
                if not is_numeric(Information_extraidos['Span atual A']):
                    avisos.append(f"{Information_extraidos['Span atual A']}")
                if not is_numeric(Information_extraidos['Span atual B']):
                    avisos.append(f"{Information_extraidos['Span atual B']}")

                if not avisos:
                    df_input.at[index, 'AVISOS'] = 'OK'
                else:
                    df_input.at[index, 'AVISOS'] = ', '.join(avisos)

                atualizados.append(index)
            else:
                print(f"Failed to retrieve data for designador {designador}.")
    
    return df_input, atualizados

def main():
    base_url = '//IP'
    session = requests.Session()
    
    file_path = os.path.join('Planilhas', 'OTSS.xlsx')
    
    os.makedirs(os.path.dirname(file_path), exist_ok=True)

    if not os.path.isfile(file_path):
        raise FileNotFoundError(f"O arquivo {file_path} não foi encontrado.")
    if not file_path.endswith('.xlsx'):
        raise ValueError(f"O arquivo {file_path} não é um arquivo Excel válido.")
    
    try:
        df_input = pd.read_excel(file_path, engine='openpyxl')
    except Exception as e:
        raise ValueError(f"Erro ao abrir o arquivo {file_path}. Verifique se é um arquivo Excel válido e não está corrompido.") from e
    
    print("Colunas do DataFrame:")
    print(df_input.columns)
    
    required_columns = ['OTS', 'TX A', 'RX A', 'Span atual A', 'Span projetado A', 'Histórico A',
                        'Equipamento de Envio (A)', 'TX B', 'RX B', 'Span atual B', 'Span projetado B',
                        'Histórico B', 'Equipamento de Recepção (B)', 'KM: A <> B', 'Status A', 'Status B', 'Trat Niveis A', 'Trat Niveis B', 'DB/KM A', 'DB/KM B', 'ATUALIZADO EM', 'Consulta URL', 'Consulta Sucesso']
    for col in required_columns:
        if col not in df_input.columns:
            df_input[col] = ''
    
    print(f"Total de linhas no DataFrame: {len(df_input)}")
    

    option = input("Escolha a ação:\n1. Atualizar Information específicos\n2. Atualizar toda a planilha\n3. Atualizar apenas novos Information\nDigite o número da opção desejada: ")

    if option == '1':
        print("Designadores disponíveis para atualização:")
        print(df_input['OTS'].tolist())
        designadores = input("Digite os designadores (OTS) a serem atualizados, separados por vírgula: ").split(',')
        designadores = [d.strip() for d in designadores]
        df_input, atualizados = atualizar_planilha(df_input, session, base_url, designadores=designadores)
    elif option == '2':
        df_input, atualizados = atualizar_planilha(df_input, session, base_url)
    elif option == '3':
        df_input, atualizados = atualizar_planilha(df_input, session, base_url, novos_Information=True)
    else:
        print("Opção inválida. Saindo do programa.")
        return

    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    for index in atualizados:
        df_input.at[index, 'ATUALIZADO EM'] = timestamp

    book = load_workbook(file_path)
    sheet_name = book.sheetnames[0]
    sheet = book[sheet_name]
    
    # Clear all cells in the sheet so there are no overwrite errors
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.value = None

    for r_idx, row in df_input.iterrows():
        for c_idx, value in enumerate(row):
            sheet.cell(row=r_idx + 2, column=c_idx + 1, value=value)
    
    book.save(file_path)
    print("Consulta e gravação de resultados concluídas com sucesso.")

if __name__ == "__main__":
    main()